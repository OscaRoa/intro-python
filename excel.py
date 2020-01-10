import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# Los archivos de excel también son llamados "libros" o "Workbooks" en inglés
wb = openpyxl.load_workbook('example.xlsx')

# Cada libro tiene distintas hojas o "sheets"
print(wb.sheetnames)

# Accedemos a una hoja como si fuera un diccionario
sheet = wb["Sheet1"]

# Entre los atributos de una hoja, se encuentra su título
print(sheet.title)

# Para obtener una celda determinada, podemos también hacerlo
# como si fuera una llave de un diccionario
celda_A1 = sheet["A1"]

# Entre los atributos de una celda, se encuentra el valor de ésta
print(celda_A1.value)

# Sin embargo si tenemos una hoja de Excel con muchas columnas
# sería impráctico y difícil de acceder automáticamente a estas columnas
# por lo que al usar el método cell() nos evitamos esta problemática
celda_A1 = sheet.cell(row=1, column=1)

# La posición de la primera fila será 1 y no 0 como en las listas,
# por lo que al iterar todas las filas tenemos que empezar por 1 en
# la función range y terminar en 8, que sería el índice número 7
for i in range(1, 8):
    celda = sheet.cell(row=i, column=2)
    print(i, celda.value)

# Para saber cuántas filas y cuántas columnas tiene nuestra hoja de Excel
# llamamos a los atributos max_row y max_column
print(sheet.max_row, sheet.max_column)

# El método que importamos al inicio nos convierte un número de columna en su letra correspondiente
print(get_column_letter(900))

# Y el otro método importado hace la operación inversa, convierte una letra de columna en un número
print(column_index_from_string("AHP"))

# También podemos acceder a un rectángulo de celdas de la siguiente manera
celdas_a1_d3 = sheet["A1":"D3"]

for fila in celdas_a1_d3:
    for celda in fila:
        print(celda.coordinate, celda.value)
    print(" -- FIN DE LA FILA --")

# A una columna en particular
# Primero tenemos que convertir el atributo columnas a una lista, después ya podemos iterar
columna_B = list(sheet.columns)[1]
for celda in columna_B:
    print(celda.value)

# Y también a todas las filas de una columna
filas = sheet.rows
for fila in filas:
    for celda in fila:
        print(celda.coordinate, celda.value)
    print("-- FIN DE LA FILA --")